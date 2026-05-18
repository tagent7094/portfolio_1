"""One-off helper to inspect two xlsx files for manisha pack source-injection."""
import sys
import openpyxl

def dump(path, max_sheets=3, max_rows=6, max_cols=8):
    print(f"\n=== {path} ===")
    wb = openpyxl.load_workbook(path, data_only=True)
    print("sheets:", wb.sheetnames)
    for sn in wb.sheetnames[:max_sheets]:
        ws = wb[sn]
        print(f"\n--- sheet: {sn} ({ws.max_row}x{ws.max_column}) ---")
        for r in range(1, min(ws.max_row + 1, max_rows + 1)):
            for c in range(1, min(ws.max_column + 1, max_cols + 1)):
                v = ws.cell(r, c).value
                if v is not None:
                    vs = str(v)[:120].replace("\n", "|")
                    print(f"  [{r},{c}]: {vs}")
            print("  ---")

if __name__ == "__main__":
    for p in sys.argv[1:]:
        dump(p)
