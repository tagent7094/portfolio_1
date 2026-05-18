"""Inject source-post text from Transpose test runs.xlsx into Manisha pack.

The Manisha adapted-posts xlsx has a `Source #` column (1..6) but no
`Source Quote` column, so the pack viewer's left panel renders empty.

This script:
  1. Reads Transpose test runs.xlsx, col 12 ("Post content"), rows 2..7 → source 1..6
  2. Reads the Manisha pack xlsx
  3. Inserts a `Source Quote` column right after `Source #`
  4. Fills it per-row by looking up the source number
  5. Writes the result to OUT_PATH
"""

from pathlib import Path
import openpyxl
from copy import copy

TRANSPOSE = Path(r"C:\Users\Akshit\Downloads\Transpose test runs.xlsx")
PACK_IN = Path(r"C:\Users\Akshit\AppData\Local\Packages\CLAUDE~1\LOCALC~1\Roaming\Claude\LOCAL-~1\FA9FC1~1\9F407A~1\LO767F~1\outputs\Manisha_Adapted_Posts_v1.xlsx")
OUT_PATH = Path(r"D:\projects\personal_graph\digital-dna\data\quality-loop\manisha_batch_2026-05-18_3_with_sources.xlsx")

SOURCE_TEXT_COL = 12   # "Post content"
SOURCE_NUMBER_COL_NAME = "Source #"
INJECTED_COL_NAME = "Source Quote"


def load_source_map(transpose_path: Path) -> dict[int, str]:
    wb = openpyxl.load_workbook(transpose_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sources: dict[int, str] = {}
    # rows 2..N each map to source 1..N-1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, SOURCE_TEXT_COL).value
        if v is not None and str(v).strip():
            sources[r - 1] = str(v).strip()
    return sources


def inject(pack_in: Path, out_path: Path, sources: dict[int, str]) -> None:
    wb = openpyxl.load_workbook(pack_in)
    ws = wb["Posts"]

    # Header row — find Source # column and pick injection point after it
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if INJECTED_COL_NAME in header_row:
        print(f"[skip] '{INJECTED_COL_NAME}' already exists — overwriting in place")
        inject_col_idx = header_row.index(INJECTED_COL_NAME) + 1
    else:
        if SOURCE_NUMBER_COL_NAME not in header_row:
            raise SystemExit(f"Couldn't find '{SOURCE_NUMBER_COL_NAME}' header. Got: {header_row}")
        src_num_col_idx = header_row.index(SOURCE_NUMBER_COL_NAME) + 1
        inject_col_idx = src_num_col_idx + 1
        ws.insert_cols(inject_col_idx)
        ws.cell(1, inject_col_idx).value = INJECTED_COL_NAME
        # Copy header formatting from the col before so it looks consistent
        try:
            src_cell = ws.cell(1, src_num_col_idx)
            dst_cell = ws.cell(1, inject_col_idx)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.border = copy(src_cell.border)
        except Exception:
            pass
        ws.column_dimensions[ws.cell(1, inject_col_idx).column_letter].width = 60

    # Fill source quote per row
    src_num_col_idx = (
        [ws.cell(1, c).value for c in range(1, ws.max_column + 1)].index(SOURCE_NUMBER_COL_NAME) + 1
    )
    filled = 0
    missing = 0
    for r in range(2, ws.max_row + 1):
        src_num_val = ws.cell(r, src_num_col_idx).value
        if src_num_val is None:
            continue
        try:
            src_num = int(src_num_val)
        except (ValueError, TypeError):
            try:
                src_num = int(float(src_num_val))
            except Exception:
                continue
        text = sources.get(src_num)
        if text:
            ws.cell(r, inject_col_idx).value = text
            ws.cell(r, inject_col_idx).alignment = openpyxl.styles.Alignment(
                wrap_text=True, vertical="top",
            )
            filled += 1
        else:
            missing += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[ok] Wrote {out_path}")
    print(f"[ok] Filled {filled} rows; {missing} rows had no source mapping.")
    print(f"[ok] Total sources loaded: {len(sources)}  (source numbers: {sorted(sources)})")


if __name__ == "__main__":
    sources = load_source_map(TRANSPOSE)
    print(f"[info] Loaded {len(sources)} source posts from {TRANSPOSE.name}")
    for n, t in sources.items():
        print(f"  src #{n}: {t[:80]}...")
    inject(PACK_IN, OUT_PATH, sources)
