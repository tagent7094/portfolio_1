"""
Post-pack routes: browse and read per-founder post packs (Excel or batch JSON).

  GET   /api/admin/founders/{slug}/post-packs                -> list available dates
  GET   /api/admin/founders/{slug}/post-packs/{date}         -> parse pack -> JSON
  POST  /api/admin/founders/{slug}/post-packs/{date}/export-sheets -> create Google Sheet
  POST  /api/admin/setup-google-key                          -> one-time: upload service-account JSON
"""

from __future__ import annotations

import json
import logging
import os
import re
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Header, HTTPException, Request, UploadFile, File
from pydantic import BaseModel

from webapp.auth_routes import _require_admin
from src.auth.tokens import decode_token

router = APIRouter(prefix="/api/admin/founders", tags=["admin-packs"])
setup_router = APIRouter(prefix="/api/admin", tags=["admin-setup"])
founder_router = APIRouter(prefix="/api/founders", tags=["founder-packs"])
logger = logging.getLogger(__name__)

FOUNDERS_DIR = Path(__file__).parent.parent / "data" / "founders"


def _require_founder(request: Request, slug: str) -> None:
    if os.environ.get("TAGENT_AUTH_ENABLED", "").lower() not in ("1", "true", "yes"):
        return
    token = request.cookies.get("tagent_token", "")
    claims = decode_token(token)
    if claims and claims.get("sub") == slug:
        return
    admin_token = request.cookies.get("admin_token", "")
    admin_claims = decode_token(admin_token)
    if admin_claims and admin_claims.get("sub") == "admin":
        return
    if not claims and not admin_claims:
        raise HTTPException(status_code=401, detail="unauthenticated")
    raise HTTPException(status_code=403, detail="not authorized for this founder")


def _post_data_dir(slug: str) -> Path:
    """Resolve the post-data/ folder for a slug, tolerating mixed-case folder names."""
    if FOUNDERS_DIR.is_dir():
        for folder in FOUNDERS_DIR.iterdir():
            if not folder.is_dir():
                continue
            folder_slug = folder.name.lower().replace(" ", "_").replace("-", "_")
            if folder_slug == slug:
                return folder / "post-data"
    return FOUNDERS_DIR / slug / "post-data"


def _extract_date(filename: str) -> str:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
    return m.group(1) if m else ""


def _latest_file_for_date(post_dir: Path, date: str, ext: str = ".json") -> Path | None:
    """Find the latest file for a date (highest _N suffix = most recent run)."""
    pattern = f"*_batch_*{ext}" if ext == ".json" else f"*{ext}"
    matches = [f for f in post_dir.glob(pattern) if date in f.name]
    if ext == ".json":
        matches = [f for f in matches if "_log" not in f.name]
    if not matches:
        return None
    return max(matches, key=lambda f: f.stat().st_mtime)


# ── Batch JSON -> tabular transformer ────────────────────────────────────────

_VARIANT_HEADERS = []
for _letter in ("A", "B", "C", "D", "E"):
    _VARIANT_HEADERS += [
        f"Variant {_letter} Opening",
        f"Variant {_letter} Rewrite Type",
        f"Variant {_letter} Key Change",
        f"Variant {_letter} Expected Lift",
    ]

_BATCH_HEADERS = [
    "Row #", "Source #", "Type", "Entry Door", "Mode",
    "Final Post", "Word Count", "Voice Score", "Violations", "Mechanic", "Original Opening",
    "Final Opening", "Rating", "Recommended", "Buried Gold", "Weakness", "Versions Considered",
    *_VARIANT_HEADERS,
    "Argument", "Events Used", "Gates", "Source Post", "Convergence",
]


def _batch_json_to_tabular(data: dict) -> dict:
    """Convert nested batch JSON to the flat {readme, headers, posts} format."""
    metadata = data.get("metadata", {})

    readme = {
        "Founder": metadata.get("founder", ""),
        "Date": str(metadata.get("generated_at", ""))[:10],
        "Posts": str(metadata.get("total_posts", 0)),
        "Sources": str(metadata.get("sources_count", 0)),
        "Platform": metadata.get("platform", "linkedin"),
        "Median word count": str(metadata.get("median_word_count", "")),
        "Pack": "Batch Cowork",
    }

    posts: list[dict[str, Any]] = []
    for pack in data.get("packs", []):
        src_num = pack.get("source_number", 0)
        source_post = str(pack.get("source_post", ""))
        conv = pack.get("convergence_test", {})
        conv_str = "PASS" if conv.get("passed", True) else f"FAIL: {conv.get('recommendation', '')}"

        for post in pack.get("posts", []):
            amp = post.get("amplifier", {})
            gates = amp.get("gates", {})
            gates_str = "; ".join(
                f"{k}={'pass' if v else 'fail'}" for k, v in gates.items()
            ) if gates else ""

            events = post.get("events_used", [])
            events_str = "; ".join(events) if isinstance(events, list) else str(events)

            row: dict[str, Any] = {
                "Row #": f"{src_num}-{post.get('label', '')}",
                "Source #": src_num,
                "Type": post.get("batch", ""),
                "Entry Door": post.get("entry_door", ""),
                "Mode": post.get("mode", ""),
                "Final Post": post.get("text", ""),
                "Word Count": post.get("word_count", 0),
                "Voice Score": post.get("voice_validation", {}).get("voice_score", ""),
                "Violations": "; ".join(post.get("violations", [])),
                "Mechanic": amp.get("mechanic", ""),
                "Original Opening": amp.get("original_opening", ""),
                "Final Opening": amp.get("final_opening", ""),
                "Rating": amp.get("rating", 0),
                "Recommended": amp.get("recommended_variant", ""),
                "Buried Gold": amp.get("buried_gold", ""),
                "Weakness": amp.get("weakness", ""),
                "Versions Considered": amp.get("versions_considered", 0),
                "Argument": post.get("argument_compressed", ""),
                "Events Used": events_str,
                "Gates": gates_str,
                "Source Post": source_post,
                "Convergence": conv_str,
            }

            variants = amp.get("variants", [])
            variant_map = {v.get("variant", ""): v for v in variants if isinstance(v, dict)}
            for letter in ("A", "B", "C", "D", "E"):
                v = variant_map.get(letter, {})
                row[f"Variant {letter} Opening"] = v.get("opening", "")
                row[f"Variant {letter} Rewrite Type"] = v.get("mechanic", "")
                row[f"Variant {letter} Key Change"] = v.get("key_change", "")
                row[f"Variant {letter} Expected Lift"] = v.get("expected_lift", "")

            posts.append(row)

    return {"readme": readme, "headers": list(_BATCH_HEADERS), "posts": posts}


def _load_pack_data(slug: str, date: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Load pack data (headers + rows) from JSON or Excel. Returns (headers, rows)."""
    post_dir = _post_data_dir(slug)

    # Try batch JSON first (latest file for this date)
    f = _latest_file_for_date(post_dir, date, ".json")
    if f:
        raw = json.loads(f.read_text(encoding="utf-8"))
        tabular = _batch_json_to_tabular(raw)
        return tabular["headers"], tabular["posts"]

    # Fall back to Excel
    target = _latest_file_for_date(post_dir, date, ".xlsx")
    if not target:
        raise HTTPException(status_code=404, detail=f"No pack found for {date}")

    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    try:
        wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not open Excel: {exc}")

    headers: list[str] = []
    rows: list[dict[str, Any]] = []
    if "Posts" in wb.sheetnames:
        ws = wb["Posts"]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
            else:
                if all(v is None for v in row):
                    continue
                post: dict[str, Any] = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        if val is None:
                            post[headers[j]] = ""
                        elif isinstance(val, float) and val == int(val):
                            post[headers[j]] = int(val)
                        else:
                            post[headers[j]] = val
                rows.append(post)
    wb.close()
    return headers, rows


# ── Shared data helpers ──────────────────────────────────────────────────────


def _list_packs_response(slug: str) -> dict:
    post_dir = _post_data_dir(slug)
    if not post_dir.exists():
        return {"packs": []}
    packs: list[dict] = []
    for f in sorted(post_dir.glob("*_batch_*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
        if "_log" in f.name:
            continue
        d = _extract_date(f.name)
        if d:
            packs.append({
                "filename": f.name, "date": d,
                "size_kb": round(f.stat().st_size / 1024, 1), "format": "json",
            })
    json_stems = {f.stem for f in post_dir.glob("*_batch_*.json")}
    for f in sorted(post_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
        d = _extract_date(f.name)
        if d and f.stem not in json_stems:
            packs.append({
                "filename": f.name, "date": d,
                "size_kb": round(f.stat().st_size / 1024, 1), "format": "xlsx",
            })
    return {"packs": packs}


def _get_pack_response(slug: str, date: str, filename: str | None = None) -> dict:
    post_dir = _post_data_dir(slug)
    if filename:
        target = post_dir / filename
        if target.exists() and target.suffix == ".json":
            raw = json.loads(target.read_text(encoding="utf-8"))
            return _batch_json_to_tabular(raw)
    f = _latest_file_for_date(post_dir, date, ".json")
    if f:
        raw = json.loads(f.read_text(encoding="utf-8"))
        return _batch_json_to_tabular(raw)
    target = _latest_file_for_date(post_dir, date, ".xlsx")
    if not target:
        raise HTTPException(status_code=404, detail=f"No pack found for {date}")
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")
    try:
        wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Could not open Excel: {exc}")
    result: dict = {"readme": {}, "headers": [], "posts": []}
    if "README" in wb.sheetnames:
        for row in wb["README"].iter_rows(values_only=True):
            if row[0] is not None and row[1] is not None:
                result["readme"][str(row[0])] = str(row[1])
    if "Posts" in wb.sheetnames:
        ws = wb["Posts"]
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                result["headers"] = headers
            else:
                if all(v is None for v in row):
                    continue
                post: dict = {}
                for j, val in enumerate(row):
                    if j < len(headers):
                        if val is None:
                            post[headers[j]] = ""
                        elif isinstance(val, float) and val == int(val):
                            post[headers[j]] = int(val)
                        else:
                            post[headers[j]] = val
                result["posts"].append(post)
    wb.close()
    return result


def _get_traces_response(slug: str, date: str) -> dict:
    post_dir = _post_data_dir(slug)
    log_matches = [f for f in post_dir.glob("*_batch_*_log.json") if date in f.name]
    if log_matches:
        log_file = max(log_matches, key=lambda f: f.stat().st_mtime)
        raw = json.loads(log_file.read_text(encoding="utf-8"))
        return {
            "traceability": raw.get("traces", raw.get("traceability", {})),
            "web_search": raw.get("web_searches", raw.get("web_search", {})),
            "pipeline_log": raw.get("pipeline_log", []),
        }
    f = _latest_file_for_date(post_dir, date, ".json")
    if f:
        raw = json.loads(f.read_text(encoding="utf-8"))
        return {
            "traceability": raw.get("traceability", {}),
            "web_search": raw.get("web_search", {}),
        }
    raise HTTPException(status_code=404, detail=f"No batch pack with traces for {date}")


# ── Admin pack endpoints ─────────────────────────────────────────────────────


@router.get("/{slug}/post-packs")
async def list_post_packs(slug: str, request: Request):
    _require_admin(request)
    return _list_packs_response(slug)


@router.get("/{slug}/post-packs/{date}")
async def get_post_pack(slug: str, date: str, request: Request, filename: str | None = None):
    _require_admin(request)
    return _get_pack_response(slug, date, filename)


@router.get("/{slug}/post-packs/{date}/traces")
async def get_pack_traces(slug: str, date: str, request: Request):
    _require_admin(request)
    return _get_traces_response(slug, date)


# ── Founder pack endpoints (tagent_token auth) ──────────────────────────────


@founder_router.get("/{slug}/post-packs")
async def founder_list_post_packs(slug: str, request: Request):
    _require_founder(request, slug)
    return _list_packs_response(slug)


@founder_router.get("/{slug}/post-packs/{date}")
async def founder_get_post_pack(slug: str, date: str, request: Request, filename: str | None = None):
    _require_founder(request, slug)
    return _get_pack_response(slug, date, filename)


@founder_router.get("/{slug}/post-packs/{date}/traces")
async def founder_get_pack_traces(slug: str, date: str, request: Request):
    _require_founder(request, slug)
    return _get_traces_response(slug, date)


# ── Feedback ─────────────────────────────────────────────────────────────────

def _feedback_path(slug: str, date: str) -> Path:
    post_dir = _post_data_dir(slug)
    return post_dir / f"feedback_{date}.json"


class FeedbackUpdate(BaseModel):
    row_id: str
    pre_feedback: str | None = None
    post_feedback: str | None = None


@founder_router.get("/{slug}/post-packs/{date}/feedback")
async def get_feedback(slug: str, date: str, request: Request):
    _require_founder(request, slug)
    fb_path = _feedback_path(slug, date)
    if fb_path.exists():
        try:
            return json.loads(fb_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


@founder_router.post("/{slug}/post-packs/{date}/feedback")
async def save_feedback(slug: str, date: str, body: FeedbackUpdate, request: Request):
    _require_founder(request, slug)
    fb_path = _feedback_path(slug, date)
    data: dict = {}
    if fb_path.exists():
        try:
            data = json.loads(fb_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    entry = data.get(body.row_id, {})
    if body.pre_feedback is not None:
        entry["pre_feedback"] = body.pre_feedback
    if body.post_feedback is not None:
        entry["post_feedback"] = body.post_feedback
    data[body.row_id] = entry
    fb_path.parent.mkdir(parents=True, exist_ok=True)
    fb_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return {"ok": True}


# ── Google Sheets export ──────────────────────────────────────────────────────

SA_FILE = Path("/etc/tagent/google-sa.json")
SHARE_WITH = "content@tagent.club"


class SheetsExportRequest(BaseModel):
    edits: dict[str, dict[str, str]] = {}


@router.post("/{slug}/post-packs/{date}/export-sheets")
async def export_to_sheets(slug: str, date: str, body: SheetsExportRequest, request: Request):
    _require_admin(request)

    if not SA_FILE.exists():
        raise HTTPException(
            status_code=503,
            detail="Google Sheets not configured. Run deploy/setup_google.sh on the VPS first.",
        )

    try:
        from google.oauth2 import service_account  # type: ignore
        from googleapiclient.discovery import build  # type: ignore
    except ImportError:
        raise HTTPException(status_code=503, detail="google-api-python-client not installed. Run: pip install google-auth google-api-python-client")

    headers, rows = _load_pack_data(slug, date)

    # Merge edits
    for row_dict in rows:
        row_id = str(row_dict.get("Row #", ""))
        if row_id in body.edits:
            for col_key, val in body.edits[row_id].items():
                if col_key in row_dict:
                    row_dict[col_key] = val

    # Build sheet values
    values = [headers] + [[str(r.get(h, "") or "") for h in headers] for r in rows]

    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        creds = service_account.Credentials.from_service_account_file(str(SA_FILE), scopes=scopes)
        sheets_svc = build("sheets", "v4", credentials=creds)
        drive_svc  = build("drive",  "v3", credentials=creds)

        title = f"{slug.replace('_', ' ').title()} -- {date}"
        spreadsheet = sheets_svc.spreadsheets().create(body={
            "properties": {"title": title},
            "sheets": [{"properties": {"title": "Posts"}}],
        }).execute()
        sid = spreadsheet["spreadsheetId"]

        sheets_svc.spreadsheets().values().update(
            spreadsheetId=sid,
            range="Posts!A1",
            valueInputOption="RAW",
            body={"values": values},
        ).execute()

        sheets_svc.spreadsheets().batchUpdate(spreadsheetId=sid, body={"requests": [
            {"repeatCell": {"range": {"sheetId": 0, "startRowIndex": 0, "endRowIndex": 1},
                            "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
                            "fields": "userEnteredFormat.textFormat.bold"}},
            {"updateSheetProperties": {"properties": {"sheetId": 0, "gridProperties": {"frozenRowCount": 1}},
                                       "fields": "gridProperties.frozenRowCount"}},
        ]}).execute()

        drive_svc.permissions().create(
            fileId=sid,
            body={"type": "user", "role": "writer", "emailAddress": SHARE_WITH},
            sendNotificationEmail=False,
        ).execute()

        url = f"https://docs.google.com/spreadsheets/d/{sid}"
        logger.info("Created Google Sheet %s for %s/%s", url, slug, date)
        return {"url": url}

    except Exception as exc:
        logger.exception("Google Sheets export failed: %s", exc)
        raise HTTPException(status_code=500, detail=f"Sheets export failed: {exc}")


# ── One-time Google service account key upload ────────────────────────────────

@setup_router.post("/setup-google-key")
async def setup_google_key(request: Request, x_setup_token: str = Header(default="")):
    expected = os.environ.get("TAGENT_SETUP_TOKEN", "")
    if not expected or x_setup_token != expected:
        raise HTTPException(status_code=403, detail="invalid setup token")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="body must be valid JSON")

    if "type" not in body or body.get("type") != "service_account":
        raise HTTPException(status_code=400, detail="body must be a service_account JSON key")

    SA_FILE.parent.mkdir(parents=True, exist_ok=True)
    SA_FILE.write_text(json.dumps(body, indent=2))
    SA_FILE.chmod(0o600)
    try:
        import pwd
        uid = pwd.getpwnam("tagent").pw_uid
        gid = pwd.getpwnam("tagent").pw_gid
        os.chown(SA_FILE, uid, gid)
    except Exception:
        pass

    logger.info("Google service account key installed at %s", SA_FILE)
    return {"ok": True, "path": str(SA_FILE)}


# ── Excel upload ────────────────────────────────────────────────────────────

@router.post("/{slug}/upload-pack")
async def upload_pack(slug: str, request: Request, file: UploadFile = File(...)):
    """Upload an Excel file as a new post pack."""
    _require_admin(request)

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    post_dir = _post_data_dir(slug)
    post_dir.mkdir(parents=True, exist_ok=True)

    from datetime import date as _date
    today = _date.today().isoformat()

    existing = [f for f in post_dir.glob(f"*{today}*.xlsx") if "upload" in f.name]
    suffix = f"_{len(existing)}" if existing else ""
    dest_name = f"{slug}_upload_{today}{suffix}.xlsx"
    dest = post_dir / dest_name

    content = await file.read()
    dest.write_bytes(content)
    logger.info("[upload] Saved %s (%d bytes)", dest, len(content))

    d = _extract_date(dest_name)
    return {"filename": dest_name, "date": d, "size_kb": round(len(content) / 1024, 1)}
