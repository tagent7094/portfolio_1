"""Standalone JSON-to-Excel converter for batch output files.

Produces a 3-sheet workbook from a batch JSON:
  README   — pack metadata (founder, date, posts, sources, platform)
  Posts    — the canonical 42-column post-pack schema (from `tabularize`)
  Run Cost — cost telemetry (total / by_task / by_model / by_pack)

Schema lives in `src/batch/tabularize.py` — DO NOT redefine columns here.

Usage:
    python -m src.batch.json_to_excel <input.json> [output.xlsx]
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from .tabularize import BATCH_HEADERS, LONG_TEXT_COLUMNS, to_readme, to_rows


def _auto_column_widths(
    rows: list[list],
    headers: list[str],
    min_w: int = 10,
    max_w: int = 90,
) -> dict[int, int]:
    """Width per column based on the longest single line of header + cell content.

    Single-line measurement (not full cell length) prevents multi-paragraph
    fields like "Final Post" from blowing up column widths.
    """
    widths: dict[int, int] = {}
    for i, h in enumerate(headers):
        widths[i] = max(min_w, min(max_w, len(str(h)) + 2))
    for row in rows:
        for i, v in enumerate(row):
            if i >= len(headers):
                continue
            text = str(v) if v is not None else ""
            longest_line = max((len(line) for line in text.splitlines()), default=len(text))
            widths[i] = max(widths[i], min(max_w, longest_line + 2))
    return widths


def convert(json_path: str, output_path: str | None = None) -> str:
    """Convert a batch output JSON to Excel. Returns the Excel path.

    Args:
        json_path: Source batch JSON file.
        output_path: Optional explicit destination. Defaults to input path with
            ".xlsx" suffix. Use this to keep the JSON / Excel / log stems in
            lockstep when called from compiler.save_output.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    path = Path(json_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {json_path}")

    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # ── README sheet ───────────────────────────────────────────────────────
    readme = to_readme(data)
    rs = wb.active
    rs.title = "README"
    rs.append(["Field", "Value"])
    for k, v in readme.items():
        rs.append([k, v])
    for cell in rs[1]:
        cell.font = Font(bold=True)
    rs.column_dimensions["A"].width = 24
    rs.column_dimensions["B"].width = 48

    # ── Posts sheet ────────────────────────────────────────────────────────
    rows = to_rows(data)
    ps = wb.create_sheet("Posts")
    ps.append(list(BATCH_HEADERS))

    long_idx = {BATCH_HEADERS.index(h) + 1 for h in LONG_TEXT_COLUMNS if h in BATCH_HEADERS}
    for row in rows:
        ps.append([row.get(h, "") for h in BATCH_HEADERS])

    # Header formatting + freeze pane
    for cell in ps[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")
    ps.freeze_panes = "A2"

    # Wrap long-content cells
    for excel_row in ps.iter_rows(min_row=2):
        for cell in excel_row:
            if cell.column in long_idx:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-size columns from content
    row_values = [[row.get(h, "") for h in BATCH_HEADERS] for row in rows]
    widths = _auto_column_widths(row_values, list(BATCH_HEADERS))
    for i, w in widths.items():
        ps.column_dimensions[get_column_letter(i + 1)].width = w

    # ── Run Cost sheet ─────────────────────────────────────────────────────
    cost = data.get("cost") or {}
    cs = wb.create_sheet("Run Cost")
    cs.append(["Metric", "Value"])
    cs.append(["Total USD", f"${float(cost.get('total_usd', 0) or 0):.4f}"])
    cs.append(["Input tokens", cost.get("total_input_tokens", 0)])
    cs.append(["Output tokens", cost.get("total_output_tokens", 0)])
    if cost.get("warning"):
        cs.append(["⚠ Warning", cost["warning"]])

    cs.append([])
    cs.append(["Cost by task (USD)", ""])
    for k, v in sorted((cost.get("by_task") or {}).items(), key=lambda x: -x[1]):
        cs.append([k, f"${float(v):.4f}"])

    cs.append([])
    cs.append(["Cost by model (USD)", ""])
    for k, v in sorted((cost.get("by_model") or {}).items(), key=lambda x: -x[1]):
        cs.append([k, f"${float(v):.4f}"])

    cs.append([])
    cs.append(["Cost by source pack (USD)", ""])
    for k, v in sorted((cost.get("by_pack") or {}).items()):
        cs.append([f"pack {k}", f"${float(v):.4f}"])

    for cell in cs[1]:
        cell.font = Font(bold=True)
    cs.column_dimensions["A"].width = 36
    cs.column_dimensions["B"].width = 24

    # ── Save ───────────────────────────────────────────────────────────────
    out = Path(output_path) if output_path else path.with_suffix(".xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(
            "Usage: python -m src.batch.json_to_excel <input.json> [output.xlsx]",
            file=sys.stderr,
        )
        sys.exit(1)

    out_arg = sys.argv[2] if len(sys.argv) >= 3 else None
    result = convert(sys.argv[1], output_path=out_arg)
    print(f"Excel saved: {result}")
