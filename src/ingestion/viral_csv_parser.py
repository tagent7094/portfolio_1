"""Parse viral LinkedIn posts CSV into structured records."""

from __future__ import annotations

import csv
import hashlib
import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)


def _parse_int(val: str) -> int:
    """Parse an integer from potentially messy CSV data."""
    if not val:
        return 0
    val = re.sub(r"[,%\s]", "", str(val))
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _parse_float(val: str) -> float:
    if not val:
        return 0.0
    val = re.sub(r"[%\s]", "", str(val))
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def parse_viral_csv(csv_path: str | Path) -> list[dict]:
    """Parse the viral LinkedIn posts CSV into structured records.

    CSV columns: LinkedIn Profile of Creator, Comments (Nitesh / Vaishali),
    Number of followers, Post URL, Content type, Recent, Applicable for,
    Likes vs followers ratio, Likes, Comments, Reposts, Post content

    Returns list of dicts with: post_id, content, likes, comments, reposts,
    followers, likes_ratio, engagement_score, content_type, creator_url
    """
    path = Path(csv_path)
    if not path.exists():
        logger.error("CSV/XLSX not found: %s", path)
        return []

    # Handle Excel files via pandas
    if path.suffix.lower() in ('.xlsx', '.xls'):
        return _parse_xlsx(path)

    records = []
    with open(path, encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return []

        for row_num, row in enumerate(reader, start=2):
            if len(row) < 12:
                continue

            content = row[11].strip() if len(row) > 11 else ""
            if not content or len(content) < 50:
                continue

            likes = _parse_int(row[8])
            comments = _parse_int(row[9])
            reposts = _parse_int(row[10])
            followers = _parse_int(row[2])
            likes_ratio = _parse_float(row[7])

            # Engagement score: weighted sum
            engagement = likes + (comments * 3) + (reposts * 2)

            post_id = hashlib.md5(content[:200].encode()).hexdigest()[:12]

            records.append({
                "post_id": post_id,
                "content": content,
                "likes": likes,
                "comments": comments,
                "reposts": reposts,
                "followers": followers,
                "likes_ratio": likes_ratio,
                "engagement_score": engagement,
                "content_type": row[4].strip() if len(row) > 4 else "",
                "creator_url": row[0].strip() if row[0] else "",
                "source_sheet": "",
            })

    logger.info("Parsed %d posts from %s", len(records), path)
    return records


def _parse_xlsx(path: Path) -> list[dict]:
    """Parse an Excel file — reads ALL sheets that contain post content."""
    import openpyxl

    logger.info("Parsing Excel file (all sheets): %s", path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    CONTENT_ALIASES = {"post content", "content"}
    LIKES_ALIASES = {"likes"}
    COMMENTS_ALIASES = {"comments"}
    REPOSTS_ALIASES = {"reposts"}
    FOLLOWERS_ALIASES = {"number of followers", "followers"}
    RATIO_ALIASES = {"likes vs followers ratio", "ltf", "likes_ratio"}
    TYPE_ALIASES = {"content type", "content_type"}
    CREATOR_ALIASES = {"linkedin profile of creator", "linkedin profile", "creator_url", "profile"}

    def _find_col(headers_lower, aliases):
        for i, h in enumerate(headers_lower):
            if h in aliases:
                return i
        return None

    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            continue

        headers_lower = [str(h).strip().lower() if h else "" for h in header_row]
        content_col = _find_col(headers_lower, CONTENT_ALIASES)
        if content_col is None:
            logger.debug("Sheet %r has no content column, skipping", sheet_name)
            continue

        likes_col = _find_col(headers_lower, LIKES_ALIASES)
        comments_col = _find_col(headers_lower, COMMENTS_ALIASES)
        reposts_col = _find_col(headers_lower, REPOSTS_ALIASES)
        followers_col = _find_col(headers_lower, FOLLOWERS_ALIASES)
        ratio_col = _find_col(headers_lower, RATIO_ALIASES)
        type_col = _find_col(headers_lower, TYPE_ALIASES)
        creator_col = _find_col(headers_lower, CREATOR_ALIASES)

        sheet_count = 0
        for row in rows_iter:
            content = str(row[content_col] or "").strip() if content_col < len(row) else ""
            if not content or len(content) < 50:
                continue

            likes = _parse_int(str(row[likes_col])) if likes_col is not None and likes_col < len(row) else 0
            comments = _parse_int(str(row[comments_col])) if comments_col is not None and comments_col < len(row) else 0
            reposts = _parse_int(str(row[reposts_col])) if reposts_col is not None and reposts_col < len(row) else 0
            followers = _parse_int(str(row[followers_col])) if followers_col is not None and followers_col < len(row) else 0
            likes_ratio = _parse_float(str(row[ratio_col])) if ratio_col is not None and ratio_col < len(row) else 0.0
            engagement = likes + (comments * 3) + (reposts * 2)
            post_id = hashlib.md5(content[:200].encode()).hexdigest()[:12]

            records.append({
                "post_id": post_id,
                "content": content,
                "likes": likes,
                "comments": comments,
                "reposts": reposts,
                "followers": followers,
                "likes_ratio": likes_ratio,
                "engagement_score": engagement,
                "content_type": str(row[type_col]).strip() if type_col is not None and type_col < len(row) and row[type_col] else "",
                "creator_url": str(row[creator_col]).strip() if creator_col is not None and creator_col < len(row) and row[creator_col] else "",
                "source_sheet": sheet_name,
            })
            sheet_count += 1

        logger.info("Sheet %r: %d posts", sheet_name, sheet_count)

    wb.close()
    logger.info("Parsed %d posts total from Excel %s", len(records), path)
    return records


def get_top_posts(records: list[dict], top_pct: float = 0.05) -> list[dict]:
    """Return top N% posts by engagement score."""
    sorted_records = sorted(records, key=lambda r: r["engagement_score"], reverse=True)
    n = max(1, int(len(sorted_records) * top_pct))
    return sorted_records[:n]
