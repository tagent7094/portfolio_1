"""Schema-agnostic XLSX walker — every sheet, headers + rows as dicts.

Distinct from viral_csv_parser._parse_xlsx which is viral-schema-specific
(column-aliased to likes/comments/reposts). This reader makes no assumptions
about column names; callers map to their own schema downstream.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TypedDict

logger = logging.getLogger(__name__)


class SheetData(TypedDict):
    sheet: str
    headers: list[str]
    rows: list[dict]
    row_count: int


def read_xlsx(path: str | Path) -> list[SheetData]:
    """Read every sheet in a workbook as headered tables.

    Each entry has the sheet name, the header row (string-coerced), and rows
    as dicts keyed by header. Empty sheets and header-only sheets return
    `row_count=0`. Returns `[]` if the file can't be opened — failures are
    logged, never raised.
    """
    p = Path(path)
    if not p.exists():
        logger.warning("[xlsx_reader] file not found: %s", p)
        return []

    try:
        import openpyxl  # type: ignore
    except ImportError as e:
        logger.error("[xlsx_reader] openpyxl not installed: %s", e)
        return []

    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    except Exception as e:
        logger.warning("[xlsx_reader] failed to open %s — %s: %s", p.name, type(e).__name__, e)
        return []

    out: list[SheetData] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                out.append({"sheet": sheet_name, "headers": [], "rows": [], "row_count": 0})
                continue

            headers = [
                (str(h).strip() if h is not None else f"col_{i}")
                for i, h in enumerate(header_row)
            ]
            seen = {}
            for i, h in enumerate(headers):
                if h in seen:
                    seen[h] += 1
                    headers[i] = f"{h}__{seen[h]}"
                else:
                    seen[h] = 0

            rows: list[dict] = []
            for row in rows_iter:
                if row is None:
                    continue
                if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                    continue
                record = {}
                for i, val in enumerate(row):
                    key = headers[i] if i < len(headers) else f"col_{i}"
                    record[key] = val
                rows.append(record)

            out.append({"sheet": sheet_name, "headers": headers, "rows": rows, "row_count": len(rows)})
    finally:
        wb.close()

    return out
